"""
Sample.py - Main Script for KNN-based Missing Value Prediction
Uses spatial-temporal voxelization and KNN with Inverse Distance Weighting

Modules:
    - knn_utils.py: Distance functions and KNN prediction
    - visualization.py: 3D plotting and error analysis
    - voxelization.py: Voxelization and data loss simulation
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Import custom modules
from knn_utils import knn_predict
from visualization import plot_3d_voxels, plot_predicted_vs_actual, plot_error_analysis
from voxelization import voxelize, simulate_data_loss


# ==================== CONFIGURATION ====================
DATA_FILE = r"c:\Users\LENOVO\OneDrive\Desktop\PythonProg\MCS\Sample_DS_AP.xlsx"
OUTPUT_DIR = r"c:\Users\LENOVO\OneDrive\Desktop\PythonProg\MCS"
LOSS_PROB = 0.9    # Probability of data loss (90%)
K = 7             # Number of nearest neighbors
ALPHA = 0.70        # Weight for spatial vs temporal (0.5 = equal weight)


# ==================== 1. LOAD AND PREPROCESS DATA ====================
print("=" * 60)
print("LOADING DATA")
print("=" * 60)

df = pd.read_excel(DATA_FILE)
# print(df.head(10))

# Convert string time to datetime
df['time'] = pd.to_datetime(df['time'].astype(str).str.strip(), format='%H:%M:%S')

# Convert to seconds and time slots
df['time_sec'] = df['time'].dt.hour * 3600 + df['time'].dt.minute * 60 + df['time'].dt.second
df['time_slot'] = df['time_sec'] // 300
print(f"\nTotal data points: {len(df)}")


# ==================== 2. VOXELIZATION (All-in-One) ====================
print("\n" + "=" * 60)
print("VOXELIZATION")
print("=" * 60)

# Use voxelize() without loss simulation first (for Graph 1)
voxels, stats = voxelize(df, simulate_loss=False)
print(f"Created {stats['total_voxels']} voxels")

   
# ==================== 3. GRAPH 1: ORIGINAL DATA ====================
print("\n" + "=" * 60)
print("GRAPH 1: Original Data (Before Loss)")
print("=" * 60)
# plot_3d_voxels(voxels, 'Graph 1: Original Data (Before Loss)', color_by='value')


# ==================== 4. SIMULATE DATA LOSS ====================
print("\n" + "=" * 60)
print(f"SIMULATING DATA LOSS ({LOSS_PROB*100:.0f}%)")
print("=" * 60)

voxels, (known_count, missing_count) = simulate_data_loss(voxels, LOSS_PROB)
print(f"Known values: {known_count}")
print(f"Missing values: {missing_count}")


# ==================== 5. GRAPH 2: AFTER DATA LOSS ====================
print("\n" + "=" * 60)
print("GRAPH 2: After Data Loss")
print("=" * 60)
# plot_3d_voxels(voxels, f'Graph 2: After {LOSS_PROB*100:.0f}% Data Loss', show_missing=True)


# ==================== 6. KNN PREDICTION ====================
print("\n" + "=" * 60)
print(f"KNN PREDICTION (K={K}, Alpha={ALPHA})")
print("=" * 60)

for v in voxels:
    if v['value'] is None:
        v['predicted'] = knn_predict(v, voxels, K, ALPHA, 'ST', 'weighted')

predicted_count = sum(1 for v in voxels if 'predicted' in v)
print(f"Predicted {predicted_count} missing values")


# ==================== 7. EVALUATE ERROR ====================
print("\n" + "=" * 60)
print("ERROR EVALUATION")
print("=" * 60)

num = sum(abs(v['true_value'] - v['predicted']) for v in voxels if 'predicted' in v)
den = sum(abs(v['true_value']) for v in voxels if 'predicted' in v)

NE = num / den if den != 0 else 0
print(f"Normalized Error: {NE*100:.2f}%")
print(f"Prediction Accuracy: {(1-NE)*100:.2f}%")


# ==================== 8. GRAPH 3: PREDICTED VS ACTUAL ====================
print("\n" + "=" * 60)
print("GRAPH 3: Predicted vs Actual Values")
print("=" * 60)
# plot_predicted_vs_actual(voxels, NE)


# ==================== 9. GRAPH 4: ERROR ANALYSIS ====================
print("\n" + "=" * 60)
print("GRAPH 4: Error Analysis")
print("=" * 60)
# plot_error_analysis(voxels)


# ==================== 10. EXPORT RESULTS TO EXCEL ====================
print("\n" + "=" * 60)
print("EXPORTING RESULTS TO EXCEL")
print("=" * 60)

wb = Workbook()

# ---- Styling ----
header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')


def style_header(ws, row, num_cols):
    """Apply header styling to a row"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


# ==================== SHEET 1: DATASET ====================
ws_data = wb.active
ws_data.title = "Dataset"

# Headers
data_headers = ['S.No', 'X (Latitude)', 'Y (Longitude)', 'T (Time)',
                 'True Value', 'Status', 'Value After Loss', 'Predicted Value', 'Absolute Error']
for col, header in enumerate(data_headers, 1):
    ws_data.cell(row=1, column=col, value=header)
style_header(ws_data, 1, len(data_headers))

# Data rows
for i, v in enumerate(voxels, 1):
    row = i + 1
    is_missing = v['value'] is None
    predicted = v.get('predicted', None)
    abs_error = abs(v['true_value'] - predicted) if predicted is not None else None

    ws_data.cell(row=row, column=1, value=i).alignment = cell_align
    ws_data.cell(row=row, column=2, value=round(v['x'], 4)).alignment = cell_align
    ws_data.cell(row=row, column=3, value=round(v['y'], 4)).alignment = cell_align
    ws_data.cell(row=row, column=4, value=round(v['t'], 4)).alignment = cell_align
    ws_data.cell(row=row, column=5, value=round(v['true_value'], 4)).alignment = cell_align

    # Status column
    status_cell = ws_data.cell(row=row, column=6)
    if is_missing:
        status_cell.value = "Missing"
        status_cell.fill = red_fill
    else:
        status_cell.value = "Known"
        status_cell.fill = green_fill
    status_cell.alignment = cell_align

    # Value after loss
    val_cell = ws_data.cell(row=row, column=7)
    val_cell.value = round(v['value'], 4) if v['value'] is not None else "—"
    val_cell.alignment = cell_align

    # Predicted value
    pred_cell = ws_data.cell(row=row, column=8)
    pred_cell.value = round(predicted, 4) if predicted is not None else "—"
    if predicted is not None:
        pred_cell.fill = yellow_fill
    pred_cell.alignment = cell_align

    # Absolute error
    err_cell = ws_data.cell(row=row, column=9)
    err_cell.value = round(abs_error, 4) if abs_error is not None else "—"
    err_cell.alignment = cell_align

    # Apply border to all cells in this row
    for col in range(1, len(data_headers) + 1):
        ws_data.cell(row=row, column=col).border = thin_border

# Auto-adjust column widths
for col in range(1, len(data_headers) + 1):
    ws_data.column_dimensions[get_column_letter(col)].width = 18


# ==================== SHEET 2: SUMMARY ====================
ws_summary = wb.create_sheet("Summary")

# Title
title_cell = ws_summary.cell(row=1, column=1, value="KNN Prediction Summary")
title_cell.font = Font(name='Calibri', bold=True, size=16, color='4472C4')
ws_summary.merge_cells('A1:H1')
title_cell.alignment = Alignment(horizontal='center')

# --- Section 1: Current Run Config ---
ws_summary.cell(row=3, column=1, value="Current Run Configuration").font = Font(bold=True, size=13)
ws_summary.merge_cells('A3:D3')

config_headers = ['Parameter', 'Value']
for col, header in enumerate(config_headers, 1):
    ws_summary.cell(row=4, column=col, value=header)
style_header(ws_summary, 4, len(config_headers))

config_data = [
    ('Loss Probability', f'{LOSS_PROB*100:.0f}%'),
    ('K (Neighbors)', K),
    ('Alpha (Spatial Weight)', ALPHA),
    ('Total Voxels', len(voxels)),
    ('Known Values', known_count),
    ('Missing Values', missing_count),
    ('Predicted Values', predicted_count),
    ('Normalized Error (NE)', f'{NE*100:.2f}%'),
    ('Accuracy Rate', f'{(1-NE)*100:.2f}%'),
]

for i, (param, val) in enumerate(config_data, 5):
    ws_summary.cell(row=i, column=1, value=param).border = thin_border
    ws_summary.cell(row=i, column=1).alignment = cell_align
    val_cell = ws_summary.cell(row=i, column=2, value=val)
    val_cell.border = thin_border
    val_cell.alignment = cell_align
    if 'Accuracy' in param:
        val_cell.fill = green_fill
    elif 'Error' in param:
        val_cell.fill = red_fill


# --- Section 2: Parameter Comparison Table ---

comparison_row = 5 + len(config_data) + 2
ws_summary.cell(row=comparison_row, column=1, value="Parameter Comparison Table").font = Font(bold=True, size=13)
ws_summary.merge_cells(f'A{comparison_row}:H{comparison_row}')

# Comparison headers
comp_headers = ['S.No', 'Loss Prob', 'K', 'Alpha', 'Known', 'Missing', 'Normalized Error (%)', 'Accuracy Rate (%)']
header_row = comparison_row + 1
for col, header in enumerate(comp_headers, 1):
    ws_summary.cell(row=header_row, column=col, value=header)
style_header(ws_summary, header_row, len(comp_headers))

# Parameter combinations to test
param_combos = [
    (LOSS_PROB, K, ALPHA),             # Original run (already computed)
    (0.9, 4, 0.5),                     # i
    (0.8, 6, 0.6),                     # ii
    (0.8, 4, 0.4),                     # iii
    (0.8, 7, 0.4),                     # iv
    (0.8, 4, 0.7),                     # v
    (0.8, 7, 0.7),                     # vi
    (0.7, 5, 0.7),                     # vii
]

print("\n" + "=" * 60)
print("RUNNING PARAMETER COMPARISONS")
print("=" * 60)

for idx, (lp, k_val, alpha_val) in enumerate(param_combos, 1):
    data_row = header_row + idx

    if idx == 1:
        # Use already computed results for the original run
        ne_val = NE
        acc_val = 1 - NE
        kn = known_count
        ms = missing_count
    else:
        # Fresh voxels for each combo
        fresh_voxels, _ = voxelize(df, simulate_loss=False)
        fresh_voxels, (kn, ms) = simulate_data_loss(fresh_voxels, lp)

        for v in fresh_voxels:
            if v['value'] is None:
                v['predicted'] = knn_predict(v, fresh_voxels, k_val, alpha_val, 'ST', 'weighted')

        num_c = sum(abs(v['true_value'] - v['predicted']) for v in fresh_voxels if 'predicted' in v)
        den_c = sum(abs(v['true_value']) for v in fresh_voxels if 'predicted' in v)
        ne_val = num_c / den_c if den_c != 0 else 0
        acc_val = 1 - ne_val

    print(f"  [{idx}] LP={lp}, K={k_val}, Alpha={alpha_val} => NE={ne_val*100:.2f}%, Acc={acc_val*100:.2f}%")

    row_data = [idx, f'{lp*100:.0f}%', k_val, alpha_val, kn, ms,
                f'{ne_val*100:.2f}%', f'{acc_val*100:.2f}%']

    for col, val in enumerate(row_data, 1):
        cell = ws_summary.cell(row=data_row, column=col, value=val)
        cell.alignment = cell_align
        cell.border = thin_border

    # Highlight accuracy
    ws_summary.cell(row=data_row, column=8).fill = green_fill
    ws_summary.cell(row=data_row, column=7).fill = red_fill

# Column widths for summary
for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_summary.column_dimensions[col_letter].width = 22

# Save the workbook
output_file = f"{OUTPUT_DIR}\\KNN_Results_LP{int(LOSS_PROB*100)}_K{K}_A{int(ALPHA*100)}.xlsx"
wb.save(output_file)
print(f"\nResults saved to: {output_file}")


print("\n" + "=" * 60)
print("COMPLETE!")
print("=" * 60)
